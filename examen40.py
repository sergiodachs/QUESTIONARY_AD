import os
import io
import csv
import json
import math
import sqlite3
import random
import hashlib
import secrets
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional

from fastapi import FastAPI, Request, UploadFile, File, Form, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, PlainTextResponse, Response
from fastapi.staticfiles import StaticFiles
from starlette.middleware.sessions import SessionMiddleware
from openpyxl import load_workbook

# ==========================================================
# SIMULADOR ADMINISTRATIVO C1 - VERSION 8.0
# ==========================================================
# Plataforma web responsive y PWA, pensada para:
# - ordenador
# - navegador móvil
# - publicación en servidor web
# - acceso por URL
# - posible empaquetado posterior como app Android (WebView / PWA)
#
# Características principales:
# - SQLite local
# - importación masiva desde Excel/CSV
# - soporte para miles de preguntas
# - modo examen y repaso inteligente
# - favoritos
# - historial
# - estadísticas por bloque
# - revisión de respuestas
# - UI responsive
# - manifest + service worker para instalación como PWA
#
# Requisitos:
#   pip install fastapi uvicorn openpyxl python-multipart itsdangerous
#
# Ejecución:
#   uvicorn app:app --host 0.0.0.0 --port 8000 --reload
#
# Si guardas este fichero con nombre app.py:
#   abre http://127.0.0.1:8000
#
# En móvil dentro de la red local:
#   uvicorn app:app --host 0.0.0.0 --port 8000
#   abre http://IP_DEL_PC:8000
#
# ==========================================================

APP_TITLE = "Simulador Administrativo C1 v8.0"
DB_PATH = "simulador_v8.db"
SECRET_KEY = os.environ.get("SIMULADOR_SECRET_KEY", secrets.token_hex(32))
ADMIN_PASSWORD = os.environ.get("SIMULADOR_ADMIN_PASSWORD", "admin123")
DEFAULT_USER = "local_user"

app = FastAPI(title=APP_TITLE)
app.add_middleware(SessionMiddleware, secret_key=SECRET_KEY)


# ==========================================================
# BASE DE DATOS
# ==========================================================

def db_connect() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    conn = db_connect()
    cur = conn.cursor()

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS questions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            external_id TEXT,
            block TEXT NOT NULL,
            subblock TEXT,
            difficulty TEXT DEFAULT 'media',
            source TEXT,
            statement TEXT NOT NULL,
            option_a TEXT NOT NULL,
            option_b TEXT NOT NULL,
            option_c TEXT NOT NULL,
            option_d TEXT NOT NULL,
            correct_option TEXT NOT NULL,
            explanation TEXT,
            tags TEXT,
            is_active INTEGER DEFAULT 1,
            checksum TEXT UNIQUE,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS attempts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            mode TEXT NOT NULL,
            total_questions INTEGER NOT NULL,
            correct_count INTEGER NOT NULL,
            wrong_count INTEGER NOT NULL,
            blank_count INTEGER NOT NULL,
            score_15 REAL NOT NULL,
            block_filter TEXT,
            duration_seconds INTEGER,
            created_at TEXT NOT NULL
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS attempt_answers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            attempt_id INTEGER NOT NULL,
            question_id INTEGER NOT NULL,
            chosen_option TEXT,
            is_correct INTEGER NOT NULL,
            created_at TEXT NOT NULL,
            FOREIGN KEY(attempt_id) REFERENCES attempts(id)
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS favorites (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            question_id INTEGER NOT NULL,
            created_at TEXT NOT NULL,
            UNIQUE(username, question_id)
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS question_progress (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            question_id INTEGER NOT NULL,
            seen_count INTEGER DEFAULT 0,
            correct_count INTEGER DEFAULT 0,
            wrong_count INTEGER DEFAULT 0,
            last_seen_at TEXT,
            next_review_at TEXT,
            ease_factor REAL DEFAULT 2.5,
            interval_days INTEGER DEFAULT 0,
            UNIQUE(username, question_id)
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS app_settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        )
        """
    )

    conn.commit()
    conn.close()


# ==========================================================
# PREGUNTAS INICIALES
# ==========================================================
# Estructura lista para crecer hasta 10.000+ preguntas.
# Aquí dejo una base corta de ejemplo + refuerzo.
# Lo normal en v8.0 es importar Excel/CSV con miles de preguntas.
# ==========================================================

SEED_QUESTIONS = [
    {
        "external_id": "Q001",
        "block": "Procedimiento administrativo",
        "subblock": "Ley 39/2015",
        "difficulty": "media",
        "source": "Seed",
        "statement": "¿Plazo general para subsanar una solicitud incompleta?",
        "option_a": "5 días hábiles",
        "option_b": "10 días",
        "option_c": "15 días",
        "option_d": "1 mes",
        "correct_option": "b",
        "explanation": "Con carácter general, se concede un plazo de 10 días para subsanar.",
        "tags": "subsanacion,plazos,ley39"
    },
    {
        "external_id": "Q002",
        "block": "Procedimiento administrativo",
        "subblock": "Ley 39/2015",
        "difficulty": "media",
        "source": "Seed",
        "statement": "¿Qué recurso cabe normalmente contra un acto que no agota la vía administrativa?",
        "option_a": "Reposición",
        "option_b": "Alzada",
        "option_c": "Revisión de oficio",
        "option_d": "Queja",
        "correct_option": "b",
        "explanation": "Con carácter general, procede recurso de alzada frente a actos que no agotan la vía administrativa.",
        "tags": "recursos,alzada"
    },
    {
        "external_id": "Q003",
        "block": "Función pública",
        "subblock": "TREBEP",
        "difficulty": "media",
        "source": "Seed",
        "statement": "¿Cada cuánto se devenga un trienio?",
        "option_a": "Cada año",
        "option_b": "Cada dos años",
        "option_c": "Cada tres años",
        "option_d": "Cada cinco años",
        "correct_option": "c",
        "explanation": "El trienio se devenga por cada tres años de servicio.",
        "tags": "trebep,trienios"
    },
    {
        "external_id": "Q004",
        "block": "Régimen local",
        "subblock": "Órganos municipales",
        "difficulty": "media",
        "source": "Seed",
        "statement": "¿Quién preside el Pleno del Ayuntamiento?",
        "option_a": "El Secretario",
        "option_b": "El Alcalde",
        "option_c": "El concejal más antiguo",
        "option_d": "La Junta de Gobierno",
        "correct_option": "b",
        "explanation": "Con carácter general, el Pleno es presidido por la Alcaldía.",
        "tags": "pleno,alcalde,regimenlocal"
    },
    {
        "external_id": "Q005",
        "block": "Transparencia",
        "subblock": "Ley 19/2014",
        "difficulty": "media",
        "source": "Seed",
        "statement": "¿Qué es la publicidad activa?",
        "option_a": "Publicidad institucional en radio",
        "option_b": "Publicar información relevante sin necesidad de solicitud previa",
        "option_c": "Responder automáticamente todos los correos",
        "option_d": "Difundir solo acuerdos del Pleno",
        "correct_option": "b",
        "explanation": "La publicidad activa obliga a publicar información relevante sin esperar a que se solicite.",
        "tags": "transparencia,publicidadactiva"
    },
    {
        "external_id": "Q006",
        "block": "Contratación pública",
        "subblock": "LCSP",
        "difficulty": "media",
        "source": "Seed",
        "statement": "¿Pueden prorrogarse los contratos menores?",
        "option_a": "Sí, siempre",
        "option_b": "No, no pueden ser objeto de prórroga",
        "option_c": "Solo si son de servicios",
        "option_d": "Solo si no superan 3.000 euros",
        "correct_option": "b",
        "explanation": "Los contratos menores no pueden ser objeto de prórroga.",
        "tags": "contratosmenores,lcsp"
    },
]


def compute_checksum(question: Dict[str, Any]) -> str:
    base = "|".join([
        question.get("block", "").strip(),
        question.get("subblock", "").strip(),
        question.get("statement", "").strip(),
        question.get("option_a", "").strip(),
        question.get("option_b", "").strip(),
        question.get("option_c", "").strip(),
        question.get("option_d", "").strip(),
        question.get("correct_option", "").strip().lower(),
    ])
    return hashlib.sha256(base.encode("utf-8")).hexdigest()


def seed_db() -> None:
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) AS c FROM questions")
    count = cur.fetchone()["c"]
    if count > 0:
        conn.close()
        return

    now = datetime.utcnow().isoformat()
    for q in SEED_QUESTIONS:
        checksum = compute_checksum(q)
        cur.execute(
            """
            INSERT OR IGNORE INTO questions (
                external_id, block, subblock, difficulty, source,
                statement, option_a, option_b, option_c, option_d,
                correct_option, explanation, tags, is_active,
                checksum, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?, ?)
            """,
            (
                q["external_id"], q["block"], q["subblock"], q["difficulty"], q["source"],
                q["statement"], q["option_a"], q["option_b"], q["option_c"], q["option_d"],
                q["correct_option"], q["explanation"], q["tags"],
                checksum, now, now,
            ),
        )

    conn.commit()
    conn.close()


# ==========================================================
# IMPORTACIÓN MASIVA
# ==========================================================
# Formato esperado en Excel/CSV:
# external_id | block | subblock | difficulty | source |
# statement | option_a | option_b | option_c | option_d |
# correct_option | explanation | tags
#
# Columnas mínimas obligatorias:
# block, statement, option_a, option_b, option_c, option_d, correct_option
# ==========================================================

def normalize_question_row(row: Dict[str, Any]) -> Dict[str, Any]:
    data = {k.strip(): ("" if row.get(k) is None else str(row.get(k)).strip()) for k in row.keys()}
    question = {
        "external_id": data.get("external_id", ""),
        "block": data.get("block", "General") or "General",
        "subblock": data.get("subblock", ""),
        "difficulty": data.get("difficulty", "media") or "media",
        "source": data.get("source", "Importado") or "Importado",
        "statement": data.get("statement", ""),
        "option_a": data.get("option_a", ""),
        "option_b": data.get("option_b", ""),
        "option_c": data.get("option_c", ""),
        "option_d": data.get("option_d", ""),
        "correct_option": data.get("correct_option", "").lower(),
        "explanation": data.get("explanation", ""),
        "tags": data.get("tags", ""),
    }
    return question


def validate_question(q: Dict[str, Any]) -> Optional[str]:
    required = ["block", "statement", "option_a", "option_b", "option_c", "option_d", "correct_option"]
    for field in required:
        if not q.get(field):
            return f"Falta el campo obligatorio: {field}"
    if q["correct_option"] not in {"a", "b", "c", "d"}:
        return "correct_option debe ser a, b, c o d"
    return None


def import_questions_from_csv(content: bytes) -> Dict[str, Any]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return bulk_insert_questions([normalize_question_row(r) for r in reader])


def import_questions_from_xlsx(content: bytes) -> Dict[str, Any]:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"inserted": 0, "duplicates": 0, "errors": ["El fichero está vacío"]}
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    items = []
    for row in rows[1:]:
        data = {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))}
        items.append(normalize_question_row(data))
    return bulk_insert_questions(items)


def bulk_insert_questions(items: List[Dict[str, Any]]) -> Dict[str, Any]:
    conn = db_connect()
    cur = conn.cursor()
    inserted = 0
    duplicates = 0
    errors = []
    now = datetime.utcnow().isoformat()

    for idx, q in enumerate(items, start=1):
        err = validate_question(q)
        if err:
            errors.append(f"Fila {idx}: {err}")
            continue
        checksum = compute_checksum(q)
        try:
            cur.execute(
                """
                INSERT INTO questions (
                    external_id, block, subblock, difficulty, source,
                    statement, option_a, option_b, option_c, option_d,
                    correct_option, explanation, tags, is_active,
                    checksum, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?, ?)
                """,
                (
                    q["external_id"], q["block"], q["subblock"], q["difficulty"], q["source"],
                    q["statement"], q["option_a"], q["option_b"], q["option_c"], q["option_d"],
                    q["correct_option"], q["explanation"], q["tags"],
                    checksum, now, now,
                ),
            )
            inserted += 1
        except sqlite3.IntegrityError:
            duplicates += 1

    conn.commit()
    conn.close()
    return {"inserted": inserted, "duplicates": duplicates, "errors": errors}


# ==========================================================
# CONSULTAS Y LÓGICA
# ==========================================================

def query_scalar(sql: str, params: tuple = ()) -> Any:
    conn = db_connect()
    cur = conn.cursor()
    cur.execute(sql, params)
    row = cur.fetchone()
    conn.close()
    if row is None:
        return None
    return row[0]


def query_all(sql: str, params: tuple = ()) -> List[sqlite3.Row]:
    conn = db_connect()
    cur = conn.cursor()
    cur.execute(sql, params)
    rows = cur.fetchall()
    conn.close()
    return rows


def get_dashboard_stats(username: str = DEFAULT_USER) -> Dict[str, Any]:
    return {
        "total_questions": query_scalar("SELECT COUNT(*) FROM questions WHERE is_active = 1") or 0,
        "total_blocks": query_scalar("SELECT COUNT(DISTINCT block) FROM questions WHERE is_active = 1") or 0,
        "total_attempts": query_scalar("SELECT COUNT(*) FROM attempts WHERE username = ?", (username,)) or 0,
        "favorites": query_scalar("SELECT COUNT(*) FROM favorites WHERE username = ?", (username,)) or 0,
        "pending_review": query_scalar(
            "SELECT COUNT(*) FROM question_progress WHERE username = ? AND next_review_at IS NOT NULL AND next_review_at <= ?",
            (username, datetime.utcnow().isoformat()),
        ) or 0,
    }


def get_blocks() -> List[str]:
    return [r["block"] for r in query_all("SELECT DISTINCT block FROM questions WHERE is_active = 1 ORDER BY block")]


def get_questions_for_exam(
    username: str,
    mode: str,
    count: int,
    block: str = "",
    include_favorites_only: bool = False,
) -> List[Dict[str, Any]]:
    params: List[Any] = []
    sql = "SELECT q.* FROM questions q"

    if include_favorites_only:
        sql += " INNER JOIN favorites f ON f.question_id = q.id AND f.username = ?"
        params.append(username)

    if mode == "review":
        sql += " LEFT JOIN question_progress p ON p.question_id = q.id AND p.username = ?"
        params.append(username)

    where = ["q.is_active = 1"]
    if block:
        where.append("q.block = ?")
        params.append(block)

    if mode == "review":
        where.append("(p.next_review_at IS NULL OR p.next_review_at <= ?)")
        params.append(datetime.utcnow().isoformat())

    sql += " WHERE " + " AND ".join(where)

    if mode == "review":
        sql += " ORDER BY COALESCE(p.next_review_at, '1970-01-01T00:00:00') ASC, RANDOM()"
    else:
        sql += " ORDER BY RANDOM()"

    sql += " LIMIT ?"
    params.append(count)

    rows = query_all(sql, tuple(params))
    return [dict(r) for r in rows]


def save_attempt(
    username: str,
    mode: str,
    block_filter: str,
    started_at: datetime,
    answers: List[Dict[str, Any]],
) -> int:
    total = len(answers)
    correct_count = sum(1 for a in answers if a["result"] == "correct")
    wrong_count = sum(1 for a in answers if a["result"] == "wrong")
    blank_count = sum(1 for a in answers if a["result"] == "blank")
    net = correct_count - (wrong_count * 0.25)
    score_15 = round(max(0.0, (net / total) * 15), 3) if total else 0.0
    duration_seconds = int((datetime.utcnow() - started_at).total_seconds())

    conn = db_connect()
    cur = conn.cursor()
    now = datetime.utcnow().isoformat()
    cur.execute(
        """
        INSERT INTO attempts (
            username, mode, total_questions, correct_count, wrong_count,
            blank_count, score_15, block_filter, duration_seconds, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            username, mode, total, correct_count, wrong_count,
            blank_count, score_15, block_filter, duration_seconds, now,
        ),
    )
    attempt_id = cur.lastrowid

    for a in answers:
        cur.execute(
            """
            INSERT INTO attempt_answers (attempt_id, question_id, chosen_option, is_correct, created_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (attempt_id, a["question_id"], a.get("chosen_option"), 1 if a["result"] == "correct" else 0, now),
        )
        update_spaced_repetition(cur, username, a["question_id"], a["result"])

    conn.commit()
    conn.close()
    return attempt_id


def update_spaced_repetition(cur: sqlite3.Cursor, username: str, question_id: int, result: str) -> None:
    cur.execute(
        "SELECT * FROM question_progress WHERE username = ? AND question_id = ?",
        (username, question_id),
    )
    row = cur.fetchone()
    now_dt = datetime.utcnow()
    now = now_dt.isoformat()

    if row is None:
        seen_count = 0
        correct_count = 0
        wrong_count = 0
        ease_factor = 2.5
        interval_days = 0
    else:
        seen_count = row["seen_count"]
        correct_count = row["correct_count"]
        wrong_count = row["wrong_count"]
        ease_factor = row["ease_factor"]
        interval_days = row["interval_days"]

    seen_count += 1

    if result == "correct":
        correct_count += 1
        if interval_days == 0:
            interval_days = 1
        elif interval_days == 1:
            interval_days = 3
        else:
            interval_days = max(1, int(round(interval_days * ease_factor)))
        ease_factor = min(3.2, ease_factor + 0.08)
    else:
        if result == "wrong":
            wrong_count += 1
        interval_days = 1
        ease_factor = max(1.3, ease_factor - 0.2)

    next_review_at = (now_dt + timedelta(days=interval_days)).isoformat()

    if row is None:
        cur.execute(
            """
            INSERT INTO question_progress (
                username, question_id, seen_count, correct_count, wrong_count,
                last_seen_at, next_review_at, ease_factor, interval_days
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                username, question_id, seen_count, correct_count, wrong_count,
                now, next_review_at, ease_factor, interval_days,
            ),
        )
    else:
        cur.execute(
            """
            UPDATE question_progress
            SET seen_count = ?, correct_count = ?, wrong_count = ?,
                last_seen_at = ?, next_review_at = ?, ease_factor = ?, interval_days = ?
            WHERE username = ? AND question_id = ?
            """,
            (
                seen_count, correct_count, wrong_count,
                now, next_review_at, ease_factor, interval_days,
                username, question_id,
            ),
        )


def is_favorite(username: str, question_id: int) -> bool:
    count = query_scalar(
        "SELECT COUNT(*) FROM favorites WHERE username = ? AND question_id = ?",
        (username, question_id),
    )
    return bool(count)


def toggle_favorite(username: str, question_id: int) -> bool:
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM favorites WHERE username = ? AND question_id = ?", (username, question_id))
    exists = cur.fetchone()[0] > 0
    now = datetime.utcnow().isoformat()

    if exists:
        cur.execute("DELETE FROM favorites WHERE username = ? AND question_id = ?", (username, question_id))
        favorite = False
    else:
        cur.execute(
            "INSERT OR IGNORE INTO favorites (username, question_id, created_at) VALUES (?, ?, ?)",
            (username, question_id, now),
        )
        favorite = True

    conn.commit()
    conn.close()
    return favorite


# ==========================================================
# HTML
# ==========================================================

def layout(title: str, body: str, request: Request) -> HTMLResponse:
    stats = get_dashboard_stats(DEFAULT_USER)
    html = f"""
    <!DOCTYPE html>
    <html lang="es">
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
        <meta name="theme-color" content="#1e3a8a">
        <title>{title}</title>
        <link rel="manifest" href="/manifest.webmanifest">
        <style>
            :root {{
                --bg:#f3f6fb;
                --card:#ffffff;
                --text:#1f2937;
                --muted:#6b7280;
                --primary:#1e3a8a;
                --primary-2:#2563eb;
                --ok:#15803d;
                --warn:#d97706;
                --bad:#b91c1c;
                --border:#dbe3ef;
                --shadow:0 10px 30px rgba(0,0,0,.08);
                --radius:20px;
            }}
            * {{ box-sizing:border-box; }}
            body {{
                margin:0;
                background:linear-gradient(180deg,#edf4ff 0%, #f8fbff 100%);
                font-family:Arial, Helvetica, sans-serif;
                color:var(--text);
            }}
            a {{ color:inherit; text-decoration:none; }}
            .app {{ max-width:1200px; margin:0 auto; padding:18px; }}
            .topbar {{
                background:linear-gradient(135deg,#1e3a8a,#2563eb);
                color:white;
                border-radius:26px;
                padding:22px;
                box-shadow:var(--shadow);
            }}
            .title {{ font-size:28px; font-weight:700; margin:0 0 8px 0; }}
            .subtitle {{ opacity:.92; font-size:14px; }}
            .grid {{ display:grid; gap:18px; }}
            .grid-4 {{ grid-template-columns:repeat(4,1fr); }}
            .grid-3 {{ grid-template-columns:repeat(3,1fr); }}
            .grid-2 {{ grid-template-columns:repeat(2,1fr); }}
            .card {{
                background:var(--card);
                border:1px solid var(--border);
                border-radius:var(--radius);
                box-shadow:var(--shadow);
                padding:18px;
            }}
            .stat-value {{ font-size:28px; font-weight:700; margin-top:6px; }}
            .muted {{ color:var(--muted); }}
            .section-title {{ font-size:20px; font-weight:700; margin:6px 0 14px; }}
            .actions {{ display:flex; flex-wrap:wrap; gap:12px; }}
            .btn {{
                display:inline-flex;
                align-items:center;
                justify-content:center;
                padding:12px 16px;
                border-radius:14px;
                border:none;
                cursor:pointer;
                font-weight:700;
                font-size:14px;
                transition:.15s transform ease, .15s opacity ease;
            }}
            .btn:hover {{ transform:translateY(-1px); }}
            .btn-primary {{ background:var(--primary); color:white; }}
            .btn-secondary {{ background:#eef4ff; color:var(--primary); border:1px solid #cfe0ff; }}
            .btn-ok {{ background:var(--ok); color:white; }}
            .btn-warn {{ background:var(--warn); color:white; }}
            .btn-bad {{ background:var(--bad); color:white; }}
            .btn-light {{ background:white; border:1px solid var(--border); color:var(--text); }}
            form.inline {{ display:inline; }}
            label {{ display:block; font-weight:700; margin-bottom:6px; }}
            input[type=text], input[type=number], input[type=password], select, textarea {{
                width:100%;
                padding:12px 14px;
                border:1px solid var(--border);
                border-radius:12px;
                background:#fbfdff;
                font-size:14px;
            }}
            textarea {{ min-height:120px; resize:vertical; }}
            .field {{ margin-bottom:14px; }}
            .question {{ font-size:22px; line-height:1.45; font-weight:700; margin-bottom:18px; }}
            .option {{
                display:block;
                padding:14px;
                border-radius:14px;
                border:1px solid var(--border);
                margin-bottom:12px;
                background:#fcfdff;
                cursor:pointer;
            }}
            .pill {{
                display:inline-block;
                padding:6px 10px;
                border-radius:999px;
                font-size:12px;
                font-weight:700;
                background:#eef2ff;
                color:#3730a3;
                margin-right:8px;
                margin-bottom:8px;
            }}
            .result-ok {{ color:var(--ok); font-weight:700; }}
            .result-bad {{ color:var(--bad); font-weight:700; }}
            .result-warn {{ color:var(--warn); font-weight:700; }}
            .table-wrap {{ overflow:auto; }}
            table {{ width:100%; border-collapse:collapse; }}
            th, td {{ padding:10px 8px; border-bottom:1px solid var(--border); text-align:left; }}
            th {{ background:#f8fbff; position:sticky; top:0; }}
            .nav {{ display:flex; flex-wrap:wrap; gap:10px; margin-top:18px; }}
            .footer-note {{ margin-top:18px; color:var(--muted); font-size:12px; }}
            .center {{ text-align:center; }}
            .spacer {{ height:18px; }}
            .danger-box {{ background:#fff5f5; border:1px solid #fecaca; color:#991b1b; padding:12px; border-radius:14px; }}
            .ok-box {{ background:#f0fdf4; border:1px solid #bbf7d0; color:#166534; padding:12px; border-radius:14px; }}
            .progress {{ width:100%; background:#e5e7eb; height:10px; border-radius:999px; overflow:hidden; }}
            .progress > div {{ background:linear-gradient(90deg,#2563eb,#1d4ed8); height:100%; }}
            @media (max-width: 960px) {{
                .grid-4, .grid-3, .grid-2 {{ grid-template-columns:1fr; }}
                .title {{ font-size:24px; }}
                .question {{ font-size:20px; }}
            }}
        </style>
    </head>
    <body>
        <div class="app">
            <div class="topbar">
                <div class="title">{APP_TITLE}</div>
                <div class="subtitle">Preparado para escritorio, móvil, web y despliegue en servidor. Banco escalable a miles de preguntas.</div>
                <div class="nav">
                    <a class="btn btn-light" href="/">Inicio</a>
                    <a class="btn btn-light" href="/exam/setup">Examen</a>
                    <a class="btn btn-light" href="/review">Repaso</a>
                    <a class="btn btn-light" href="/favorites">Favoritas</a>
                    <a class="btn btn-light" href="/history">Historial</a>
                    <a class="btn btn-light" href="/analytics">Analítica</a>
                    <a class="btn btn-light" href="/admin">Admin</a>
                </div>
            </div>

            <div class="spacer"></div>

            <div class="grid grid-4">
                <div class="card"><div class="muted">Preguntas activas</div><div class="stat-value">{stats['total_questions']}</div></div>
                <div class="card"><div class="muted">Bloques</div><div class="stat-value">{stats['total_blocks']}</div></div>
                <div class="card"><div class="muted">Intentos</div><div class="stat-value">{stats['total_attempts']}</div></div>
                <div class="card"><div class="muted">Pendientes de repaso</div><div class="stat-value">{stats['pending_review']}</div></div>
            </div>

            <div class="spacer"></div>
            {body}
            <div class="footer-note">Consejo: para 10.000 preguntas usa importación Excel/CSV. Esta arquitectura ya está preparada para ello.</div>
        </div>

        <script>
            if ('serviceWorker' in navigator) {{
                window.addEventListener('load', function() {{
                    navigator.serviceWorker.register('/sw.js').catch(function(err) {{ console.log(err); }});
                }});
            }}
        </script>
    </body>
    </html>
    """
    return HTMLResponse(html)


# ==========================================================
# RUTAS
# ==========================================================

@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    blocks = get_blocks()
    recent_attempts = query_all(
        "SELECT * FROM attempts WHERE username = ? ORDER BY id DESC LIMIT 8",
        (DEFAULT_USER,),
    )
    body = f"""
    <div class="grid grid-2">
        <div class="card">
            <div class="section-title">Empezar</div>
            <div class="actions">
                <a class="btn btn-primary" href="/exam/setup">Nuevo examen</a>
                <a class="btn btn-secondary" href="/review">Repaso inteligente</a>
                <a class="btn btn-secondary" href="/favorites">Repasar favoritas</a>
            </div>
            <div class="spacer"></div>
            <div class="muted">Bloques disponibles:</div>
            <div style="margin-top:10px;">{''.join(f'<span class="pill">{b}</span>' for b in blocks[:20])}</div>
        </div>

        <div class="card">
            <div class="section-title">Arquitectura v8.0</div>
            <ul>
                <li>SQLite escalable a miles de preguntas</li>
                <li>Importación masiva desde Excel y CSV</li>
                <li>PWA instalable en móvil</li>
                <li>Preparado para desplegar en servidor o VPS</li>
                <li>Base lista para empaquetado Android tipo WebView</li>
            </ul>
        </div>
    </div>

    <div class="spacer"></div>
    <div class="card">
        <div class="section-title">Últimos intentos</div>
        <div class="table-wrap">
            <table>
                <thead>
                    <tr>
                        <th>Fecha</th>
                        <th>Modo</th>
                        <th>Total</th>
                        <th>Aciertos</th>
                        <th>Fallos</th>
                        <th>Blancos</th>
                        <th>Puntos / 15</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join([
                        f"<tr><td>{r['created_at'][:19].replace('T',' ')}</td><td>{r['mode']}</td><td>{r['total_questions']}</td><td>{r['correct_count']}</td><td>{r['wrong_count']}</td><td>{r['blank_count']}</td><td>{r['score_15']}</td></tr>"
                        for r in recent_attempts
                    ]) or '<tr><td colspan="7">Todavía no hay intentos guardados.</td></tr>'}
                </tbody>
            </table>
        </div>
    </div>
    """
    return layout(APP_TITLE, body, request)


@app.get("/exam/setup", response_class=HTMLResponse)
def exam_setup(request: Request):
    block_options = "".join([f'<option value="{b}">{b}</option>' for b in get_blocks()])
    body = f"""
    <div class="card">
        <div class="section-title">Configurar examen</div>
        <form method="post" action="/exam/start">
            <div class="grid grid-2">
                <div class="field">
                    <label>Bloque</label>
                    <select name="block">
                        <option value="">Todos</option>
                        {block_options}
                    </select>
                </div>
                <div class="field">
                    <label>Número de preguntas</label>
                    <input type="number" name="count" value="30" min="1" max="500">
                </div>
            </div>
            <div class="grid grid-2">
                <div class="field">
                    <label>Modo</label>
                    <select name="mode">
                        <option value="exam">Examen aleatorio</option>
                        <option value="review">Repaso inteligente</option>
                        <option value="favorites">Solo favoritas</option>
                    </select>
                </div>
                <div class="field">
                    <label>Mezclar respuestas</label>
                    <select name="shuffle_options">
                        <option value="1">Sí</option>
                        <option value="0">No</option>
                    </select>
                </div>
            </div>
            <button class="btn btn-primary" type="submit">Empezar</button>
        </form>
    </div>
    """
    return layout("Configurar examen", body, request)


@app.post("/exam/start")
def exam_start(
    request: Request,
    block: str = Form(""),
    count: int = Form(30),
    mode: str = Form("exam"),
    shuffle_options: int = Form(1),
):
    include_favorites = mode == "favorites"
    real_mode = "review" if mode == "review" else "exam"
    questions = get_questions_for_exam(DEFAULT_USER, real_mode, count, block, include_favorites_only=include_favorites)

    if not questions:
        return RedirectResponse(url="/exam/setup?empty=1", status_code=303)

    if shuffle_options:
        for q in questions:
            pairs = [("a", q["option_a"]), ("b", q["option_b"]), ("c", q["option_c"]), ("d", q["option_d"])]
            random.shuffle(pairs)
            q["shuffled_options"] = pairs
        
    else:
        for q in questions:
            q["shuffled_options"] = [("a", q["option_a"]), ("b", q["option_b"]), ("c", q["option_c"]), ("d", q["option_d"])]

    request.session["current_exam"] = {
        "started_at": datetime.utcnow().isoformat(),
        "mode": mode,
        "block": block,
        "questions": [
            {
                "id": q["id"],
                "block": q["block"],
                "subblock": q["subblock"],
                "difficulty": q["difficulty"],
                "statement": q["statement"],
                "correct_option": q["correct_option"],
                "explanation": q["explanation"],
                "shuffled_options": q["shuffled_options"],
            }
            for q in questions
        ],
    }
    return RedirectResponse(url="/exam", status_code=303)


@app.get("/exam", response_class=HTMLResponse)
def exam_page(request: Request):
    exam = request.session.get("current_exam")
    if not exam:
        return RedirectResponse(url="/exam/setup", status_code=303)

    questions = exam["questions"]
    total = len(questions)
    forms = []
    for i, q in enumerate(questions, start=1):
        options_html = ""
        for real_letter, text in q["shuffled_options"]:
            options_html += f'''
            <label class="option">
                <input type="radio" name="q_{q['id']}" value="{real_letter}"> {text}
            </label>
            '''
        forms.append(f'''
        <div class="card">
            <div><span class="pill">{q['block']}</span><span class="pill">{q['subblock'] or 'General'}</span><span class="pill">{q['difficulty']}</span></div>
            <div class="question">{i}. {q['statement']}</div>
            {options_html}
            <label class="option"><input type="radio" name="q_{q['id']}" value=""> Dejar en blanco</label>
        </div>
        ''')

    body = f"""
    <div class="card">
        <div class="section-title">Examen en curso</div>
        <div class="muted">Total de preguntas: {total}</div>
        <div class="spacer"></div>
        <div class="progress"><div style="width:100%"></div></div>
    </div>
    <div class="spacer"></div>
    <form method="post" action="/exam/submit">
        {''.join(forms)}
        <div class="spacer"></div>
        <button class="btn btn-primary" type="submit">Finalizar examen</button>
    </form>
    """
    return layout("Examen", body, request)


@app.post("/exam/submit", response_class=HTMLResponse)
async def exam_submit(request: Request):
    exam = request.session.get("current_exam")
    if not exam:
        return RedirectResponse(url="/exam/setup", status_code=303)

    form = await request.form()
    answers = []

    for q in exam["questions"]:
        chosen = str(form.get(f"q_{q['id']}", "")).strip().lower()
        if not chosen:
            result = "blank"
        elif chosen == q["correct_option"]:
            result = "correct"
        else:
            result = "wrong"
        answers.append({
            "question_id": q["id"],
            "statement": q["statement"],
            "block": q["block"],
            "subblock": q["subblock"],
            "chosen_option": chosen or None,
            "correct_option": q["correct_option"],
            "explanation": q["explanation"],
            "shuffled_options": q["shuffled_options"],
            "result": result,
        })

    attempt_id = save_attempt(
        DEFAULT_USER,
        exam["mode"],
        exam["block"],
        datetime.fromisoformat(exam["started_at"]),
        answers,
    )

    request.session["last_result"] = {
        "attempt_id": attempt_id,
        "answers": answers,
        "started_at": exam["started_at"],
        "mode": exam["mode"],
        "block": exam["block"],
    }
    request.session.pop("current_exam", None)
    return RedirectResponse(url=f"/result/{attempt_id}", status_code=303)


@app.get("/result/{attempt_id}", response_class=HTMLResponse)
def result_page(attempt_id: int, request: Request):
    result = request.session.get("last_result")
    attempt = query_all("SELECT * FROM attempts WHERE id = ?", (attempt_id,))
    if not attempt:
        raise HTTPException(status_code=404, detail="Intento no encontrado")
    a = attempt[0]

    review_btn = f'<a class="btn btn-secondary" href="/result/{attempt_id}/review">Revisar respuestas</a>'
    body = f"""
    <div class="card center">
        <div class="section-title">Resultado del examen</div>
        <div class="pill">Modo: {a['mode']}</div>
        <div class="pill">Bloque: {a['block_filter'] or 'Todos'}</div>
        <div class="spacer"></div>
        <div style="font-size:40px; font-weight:700; color:{'#15803d' if a['score_15'] >= 7.5 else '#b91c1c'}">{a['score_15']} / 15</div>
        <div class="spacer"></div>
        <div>Aciertos: <strong>{a['correct_count']}</strong></div>
        <div>Fallos: <strong>{a['wrong_count']}</strong></div>
        <div>Blancos: <strong>{a['blank_count']}</strong></div>
        <div>Duración: <strong>{a['duration_seconds']} s</strong></div>
        <div class="spacer"></div>
        <div class="actions" style="justify-content:center;">
            {review_btn}
            <a class="btn btn-primary" href="/exam/setup">Nuevo examen</a>
            <a class="btn btn-light" href="/">Inicio</a>
        </div>
    </div>
    """
    return layout("Resultado", body, request)


@app.get("/result/{attempt_id}/review", response_class=HTMLResponse)
def review_page(attempt_id: int, request: Request):
    result = request.session.get("last_result")
    if not result or result.get("attempt_id") != attempt_id:
        return RedirectResponse(url=f"/result/{attempt_id}", status_code=303)

    answers = result["answers"]
    order = {"wrong": 0, "blank": 1, "correct": 2}
    answers = sorted(answers, key=lambda x: (order[x["result"]], x["block"], x["question_id"]))

    items = []
    for idx, a in enumerate(answers, start=1):
        status_class = "result-ok" if a["result"] == "correct" else ("result-bad" if a["result"] == "wrong" else "result-warn")
        status_text = "Correcta" if a["result"] == "correct" else ("Incorrecta" if a["result"] == "wrong" else "En blanco")

        options_html = []
        for letter, text in a["shuffled_options"]:
            marker = []
            if letter == a["correct_option"]:
                marker.append("Correcta")
            if a.get("chosen_option") == letter and letter != a["correct_option"]:
                marker.append("Tu respuesta")
            suffix = f" <strong>({' / '.join(marker)})</strong>" if marker else ""
            options_html.append(f"<div>{letter.upper()}) {text}{suffix}</div>")

        fav_text = "Quitar favorita" if is_favorite(DEFAULT_USER, a['question_id']) else "Guardar como favorita"
        items.append(f"""
        <div class="card">
            <div><span class="pill">{a['block']}</span><span class="pill">{a['subblock'] or 'General'}</span></div>
            <div class="question">{idx}. {a['statement']}</div>
            <div class="{status_class}">{status_text}</div>
            <div class="spacer"></div>
            {''.join(f'<div style="margin-bottom:6px;">{x}</div>' for x in options_html)}
            <div class="spacer"></div>
            <div class="muted"><strong>Explicación:</strong> {a['explanation'] or 'Sin explicación guardada.'}</div>
            <div class="spacer"></div>
            <form class="inline" method="post" action="/favorite/toggle">
                <input type="hidden" name="question_id" value="{a['question_id']}">
                <input type="hidden" name="next_url" value="/result/{attempt_id}/review">
                <button class="btn btn-secondary" type="submit">{fav_text}</button>
            </form>
        </div>
        """)

    body = f"""
    <div class="card">
        <div class="section-title">Revisión de respuestas</div>
        <div class="muted">Ordenadas mostrando primero incorrectas y en blanco para estudiar mejor.</div>
    </div>
    <div class="spacer"></div>
    {''.join(items)}
    """
    return layout("Revisión", body, request)


@app.get("/review", response_class=HTMLResponse)
def review_setup(request: Request):
    block_options = "".join([f'<option value="{b}">{b}</option>' for b in get_blocks()])
    body = f"""
    <div class="card">
        <div class="section-title">Repaso inteligente</div>
        <form method="post" action="/exam/start">
            <input type="hidden" name="mode" value="review">
            <div class="grid grid-2">
                <div class="field">
                    <label>Bloque</label>
                    <select name="block">
                        <option value="">Todos</option>
                        {block_options}
                    </select>
                </div>
                <div class="field">
                    <label>Número de preguntas</label>
                    <input type="number" name="count" value="20" min="1" max="300">
                </div>
            </div>
            <div class="field">
                <label>Mezclar respuestas</label>
                <select name="shuffle_options">
                    <option value="1">Sí</option>
                    <option value="0">No</option>
                </select>
            </div>
            <button class="btn btn-primary" type="submit">Empezar repaso</button>
        </form>
    </div>
    """
    return layout("Repaso", body, request)


@app.get("/favorites", response_class=HTMLResponse)
def favorites_page(request: Request):
    rows = query_all(
        """
        SELECT q.* FROM favorites f
        INNER JOIN questions q ON q.id = f.question_id
        WHERE f.username = ?
        ORDER BY f.id DESC
        """,
        (DEFAULT_USER,),
    )

    items = []
    for q in rows:
        items.append(f"""
        <tr>
            <td>{q['id']}</td>
            <td>{q['block']}</td>
            <td>{q['statement']}</td>
            <td>
                <form class="inline" method="post" action="/favorite/toggle">
                    <input type="hidden" name="question_id" value="{q['id']}">
                    <input type="hidden" name="next_url" value="/favorites">
                    <button class="btn btn-bad" type="submit">Quitar</button>
                </form>
            </td>
        </tr>
        """)

    body = f"""
    <div class="card">
        <div class="section-title">Preguntas favoritas</div>
        <div class="actions">
            <form class="inline" method="post" action="/exam/start">
                <input type="hidden" name="mode" value="favorites">
                <input type="hidden" name="count" value="50">
                <input type="hidden" name="block" value="">
                <input type="hidden" name="shuffle_options" value="1">
                <button class="btn btn-primary" type="submit">Examinar favoritas</button>
            </form>
        </div>
        <div class="spacer"></div>
        <div class="table-wrap">
            <table>
                <thead>
                    <tr><th>ID</th><th>Bloque</th><th>Pregunta</th><th>Acción</th></tr>
                </thead>
                <tbody>
                    {''.join(items) or '<tr><td colspan="4">No tienes preguntas favoritas todavía.</td></tr>'}
                </tbody>
            </table>
        </div>
    </div>
    """
    return layout("Favoritas", body, request)


@app.post("/favorite/toggle")
def favorite_toggle(
    question_id: int = Form(...),
    next_url: str = Form("/favorites"),
):
    toggle_favorite(DEFAULT_USER, question_id)
    return RedirectResponse(url=next_url, status_code=303)


@app.get("/history", response_class=HTMLResponse)
def history_page(request: Request):
    rows = query_all(
        "SELECT * FROM attempts WHERE username = ? ORDER BY id DESC LIMIT 200",
        (DEFAULT_USER,),
    )
    body = f"""
    <div class="card">
        <div class="section-title">Historial</div>
        <div class="table-wrap">
            <table>
                <thead>
                    <tr>
                        <th>Fecha</th>
                        <th>Modo</th>
                        <th>Bloque</th>
                        <th>Total</th>
                        <th>Aciertos</th>
                        <th>Fallos</th>
                        <th>Blancos</th>
                        <th>Puntos / 15</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join([
                        f"<tr><td>{r['created_at'][:19].replace('T',' ')}</td><td>{r['mode']}</td><td>{r['block_filter'] or 'Todos'}</td><td>{r['total_questions']}</td><td>{r['correct_count']}</td><td>{r['wrong_count']}</td><td>{r['blank_count']}</td><td>{r['score_15']}</td></tr>"
                        for r in rows
                    ]) or '<tr><td colspan="8">No hay historial todavía.</td></tr>'}
                </tbody>
            </table>
        </div>
    </div>
    """
    return layout("Historial", body, request)


@app.get("/analytics", response_class=HTMLResponse)
def analytics_page(request: Request):
    by_block = query_all(
        """
        SELECT q.block,
               COUNT(*) AS total_seen,
               SUM(CASE WHEN aa.is_correct = 1 THEN 1 ELSE 0 END) AS total_correct
        FROM attempt_answers aa
        INNER JOIN questions q ON q.id = aa.question_id
        INNER JOIN attempts a ON a.id = aa.attempt_id
        WHERE a.username = ?
        GROUP BY q.block
        ORDER BY q.block
        """,
        (DEFAULT_USER,),
    )

    worst_questions = query_all(
        """
        SELECT q.id, q.block, q.statement,
               SUM(CASE WHEN aa.is_correct = 0 THEN 1 ELSE 0 END) AS wrongs,
               COUNT(*) AS total
        FROM attempt_answers aa
        INNER JOIN questions q ON q.id = aa.question_id
        INNER JOIN attempts a ON a.id = aa.attempt_id
        WHERE a.username = ?
        GROUP BY q.id, q.block, q.statement
        HAVING COUNT(*) >= 1
        ORDER BY wrongs DESC, total DESC
        LIMIT 25
        """,
        (DEFAULT_USER,),
    )

    block_rows = []
    for r in by_block:
        total_seen = r["total_seen"] or 0
        total_correct = r["total_correct"] or 0
        ratio = round((total_correct / total_seen) * 100, 2) if total_seen else 0
        block_rows.append(f"<tr><td>{r['block']}</td><td>{total_seen}</td><td>{total_correct}</td><td>{ratio}%</td></tr>")

    wrong_rows = []
    for r in worst_questions:
        wrong_rows.append(f"<tr><td>{r['id']}</td><td>{r['block']}</td><td>{r['statement']}</td><td>{r['wrongs']}</td><td>{r['total']}</td></tr>")

    body = f"""
    <div class="grid grid-2">
        <div class="card">
            <div class="section-title">Rendimiento por bloque</div>
            <div class="table-wrap">
                <table>
                    <thead><tr><th>Bloque</th><th>Vistas</th><th>Aciertos</th><th>% acierto</th></tr></thead>
                    <tbody>{''.join(block_rows) or '<tr><td colspan="4">Sin datos todavía.</td></tr>'}</tbody>
                </table>
            </div>
        </div>
        <div class="card">
            <div class="section-title">Preguntas más falladas</div>
            <div class="table-wrap">
                <table>
                    <thead><tr><th>ID</th><th>Bloque</th><th>Pregunta</th><th>Fallos</th><th>Total</th></tr></thead>
                    <tbody>{''.join(wrong_rows) or '<tr><td colspan="5">Sin datos todavía.</td></tr>'}</tbody>
                </table>
            </div>
        </div>
    </div>
    """
    return layout("Analítica", body, request)


@app.get("/admin", response_class=HTMLResponse)
def admin_page(request: Request):
    body = f"""
    <div class="grid grid-2">
        <div class="card">
            <div class="section-title">Importar preguntas</div>
            <form method="post" action="/admin/import" enctype="multipart/form-data">
                <div class="field">
                    <label>Contraseña admin</label>
                    <input type="password" name="password" required>
                </div>
                <div class="field">
                    <label>Fichero CSV o XLSX</label>
                    <input type="file" name="file" accept=".csv,.xlsx" required>
                </div>
                <button class="btn btn-primary" type="submit">Importar</button>
            </form>
        </div>

        <div class="card">
            <div class="section-title">Formato recomendado del Excel</div>
            <div class="muted">Columnas:</div>
            <pre>external_id | block | subblock | difficulty | source | statement | option_a | option_b | option_c | option_d | correct_option | explanation | tags</pre>
            <div class="ok-box">Campos mínimos: block, statement, option_a, option_b, option_c, option_d, correct_option</div>
        </div>
    </div>

    <div class="spacer"></div>
    <div class="card">
        <div class="section-title">Operaciones de mantenimiento</div>
        <div class="actions">
            <a class="btn btn-light" href="/api/questions/export">Exportar preguntas JSON</a>
            <a class="btn btn-light" href="/api/health">Health check API</a>
        </div>
    </div>
    """
    return layout("Administración", body, request)


@app.post("/admin/import", response_class=HTMLResponse)
async def admin_import(request: Request, password: str = Form(...), file: UploadFile = File(...)):
    if password != ADMIN_PASSWORD:
        body = '<div class="danger-box">Contraseña de administración incorrecta.</div>'
        return layout("Importación", body, request)

    content = await file.read()
    filename = (file.filename or "").lower()

    try:
        if filename.endswith(".csv"):
            result = import_questions_from_csv(content)
        elif filename.endswith(".xlsx"):
            result = import_questions_from_xlsx(content)
        else:
            result = {"inserted": 0, "duplicates": 0, "errors": ["Formato no soportado. Usa CSV o XLSX."]}
    except Exception as exc:
        result = {"inserted": 0, "duplicates": 0, "errors": [str(exc)]}

    errors_html = "".join(f"<li>{e}</li>" for e in result["errors"][:50])
    body = f"""
    <div class="card">
        <div class="section-title">Resultado de la importación</div>
        <div class="ok-box">Insertadas: {result['inserted']} | Duplicadas: {result['duplicates']}</div>
        <div class="spacer"></div>
        <div class="danger-box">
            <strong>Errores detectados:</strong>
            <ul>{errors_html or '<li>Sin errores.</li>'}</ul>
        </div>
        <div class="spacer"></div>
        <a class="btn btn-primary" href="/admin">Volver a administración</a>
    </div>
    """
    return layout("Resultado importación", body, request)


# ==========================================================
# API
# ==========================================================

@app.get("/api/health")
def api_health():
    return JSONResponse({
        "ok": True,
        "app": APP_TITLE,
        "timestamp": datetime.utcnow().isoformat(),
        "total_questions": query_scalar("SELECT COUNT(*) FROM questions WHERE is_active = 1") or 0,
    })


@app.get("/api/questions/export")
def export_questions_json():
    rows = query_all("SELECT * FROM questions WHERE is_active = 1 ORDER BY id")
    payload = [dict(r) for r in rows]
    return JSONResponse(payload)


# ==========================================================
# PWA
# ==========================================================

@app.get("/manifest.webmanifest")
def manifest():
    data = {
        "name": APP_TITLE,
        "short_name": "Simulador C1",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#f3f6fb",
        "theme_color": "#1e3a8a",
        "description": "Simulador web y móvil para oposiciones administrativas C1",
        "icons": [],
    }
    return JSONResponse(data, media_type="application/manifest+json")


@app.get("/sw.js")
def service_worker():
    js = """
    const CACHE_NAME = 'simulador-c1-v8-cache';
    const URLS = ['/', '/exam/setup', '/review', '/favorites', '/history', '/analytics'];

    self.addEventListener('install', (event) => {
      event.waitUntil(caches.open(CACHE_NAME).then(cache => cache.addAll(URLS)));
    });

    self.addEventListener('fetch', (event) => {
      event.respondWith(
        caches.match(event.request).then((response) => response || fetch(event.request))
      );
    });
    """
    return Response(content=js, media_type="application/javascript")


# ==========================================================
# INICIO
# ==========================================================
init_db()
seed_db()
